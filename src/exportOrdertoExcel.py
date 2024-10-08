# -*- coding: utf-8 -*-
"""
Created on Sat Oct  5 01:40:46 2024

@author: Venkat Raja
"""

from collections import defaultdict
from datetime import datetime
import datetime
import mysql.connector as mysql
import openpyxl
import os

mydb = mysql.connect(
    host="localhost", port="3306", database="mani", user="mani", password="xxxx"
)


## Inputs
def getInput(text):
    date_input = input(text)
    try:
        user_date = datetime.datetime.strptime(date_input, "%d-%m-%Y").date()
    except ValueError:
        print("Invalid date format. Please use YYYY-MM-DD.")
        user_date = getInput(text)

    return user_date


start_date = getInput("Enter start date in DD-MM-YYYY format : ")
end_date = getInput("Enter end date in DD-MM-YYYY format : ")

cdate = datetime.datetime.now()
filename = "OrderData_" + cdate.strftime("%d-%m-%Y") + ".xlsx"

cursor = mydb.cursor()
sql_str = f"""
    SELECT indexNum, userid, brokername, brokerid, pnl, orderDate 
    FROM pnlData 
    WHERE orderDate BETWEEN '{start_date}' AND '{end_date}'
"""
cursor.execute(sql_str)
dbData = cursor.fetchall()
mydb.close()

sorted_data = sorted(dbData, key=lambda x: int(x[0]))
data_dict = defaultdict(dict)
unique_dates = set()

for row in sorted_data:
    idx, user, broker, broker_id, pnl, order_date = row
    data_dict[(idx, user, broker, broker_id)][order_date] = pnl
    unique_dates.add(order_date)

unique_dates = sorted(unique_dates)
base_columns = ["Index", "UserID", "Broker", "BrokerID"]

if os.path.exists(filename):
    wb = openpyxl.load_workbook(filename)
else:
    wb = openpyxl.Workbook()
ws = wb.active
ws.delete_rows(1, ws.max_row)

ws.append(base_columns + [dateVar.strftime("%d %b %y") for dateVar in unique_dates])

for (idx, user, broker, broker_id), pnl_data in data_dict.items():
    row_data = [idx, user, broker, broker_id]
    for dateVar in unique_dates:
        row_data.append(
            pnl_data.get(dateVar, "")
        )  # Empty string if no PnL for that date
    ws.append(row_data)

wb.save(filename)
